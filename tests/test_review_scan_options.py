from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook

from bzr_step_count.review_gui import load_gui_settings, save_gui_settings
from bzr_step_count.review_scan import (
    ReviewScanOptions,
    list_review_case_candidates,
    scan_review_root_with_options,
)


PHASES = ["外部仕様書", "内部仕様書", "コード", "テスト仕様書", "リリース後"]
HEADERS = [
    "No",
    "重大度",
    "指摘箇所",
    "指摘内容",
    "対応日時",
    "状況",
    "備考",
    "作業担当者",
    "レビュー担当者",
    "指摘分類",
    "指標対象",
    "検出工程",
    "原因工程",
]


def _write_case_workbook(path: Path, case_id: str, case_name: str, start: str, end: str) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    management = workbook.create_sheet("_集計管理")
    for row in [
        ("項目", "値"),
        ("案件ID", case_id),
        ("案件名", case_name),
        ("コード変更ステップ数", 100),
        ("外部仕様書ページ数", 10),
        ("内部仕様書ページ数", 10),
        ("テスト仕様書ページ数", 10),
        ("外部仕様書文字数", 5000),
        ("内部仕様書文字数", 3000),
        ("テスト仕様書文字数", 2000),
        ("流出不良件数", 0),
        ("レビュー開始日", start),
        ("レビュー終了日", end),
    ]:
        management.append(row)

    for phase in PHASES:
        sheet = workbook.create_sheet(phase)
        sheet.append(["指摘項目"])
        sheet.append(HEADERS)
        if phase == "外部仕様書":
            sheet.append([1, "B", "1章", f"{case_name}の指摘", "2026-05-10", "対応済", "", "佐藤", "田中", "不良", "対象", phase, phase])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def test_list_review_case_candidates_reports_metadata_and_validation_status(tmp_path):
    workbook_path = tmp_path / "CASE-001" / "レビュー結果記録表.xlsx"
    _write_case_workbook(workbook_path, "CASE-001", "注文登録", "2026-05-01", "2026-05-10")

    candidates = list_review_case_candidates(tmp_path)

    assert len(candidates) == 1
    candidate = candidates[0]
    assert candidate.case_id == "CASE-001"
    assert candidate.case_name == "注文登録"
    assert candidate.review_start == "2026-05-01"
    assert candidate.review_end == "2026-05-10"
    assert candidate.path == str(workbook_path)
    assert candidate.validation_status == "ok"


def test_scan_options_filter_by_date_range_and_selected_cases(tmp_path):
    root = tmp_path / "cases"
    case_001 = root / "CASE-001" / "レビュー結果記録表.xlsx"
    case_002 = root / "CASE-002" / "レビュー結果記録表.xlsx"
    case_003 = root / "CASE-003" / "レビュー結果記録表.xlsx"
    _write_case_workbook(case_001, "CASE-001", "注文登録", "2026-05-01", "2026-05-10")
    _write_case_workbook(case_002, "CASE-002", "在庫改善", "2026-04-01", "2026-04-10")
    _write_case_workbook(case_003, "CASE-003", "帳票出力", "2026-05-05", "2026-05-15")
    output_dir = tmp_path / "out"

    dataset = scan_review_root_with_options(
        ReviewScanOptions(
            root=root,
            output_dir=output_dir,
            start_date="2026-05-01",
            end_date="2026-05-31",
            included_workbook_paths=(str(case_001), str(case_003)),
            excluded_workbook_paths=(str(case_003),),
            skip_bazaar=True,
        )
    )

    assert [case.metadata.case_id for case in dataset.cases] == ["CASE-001"]
    rows = list(csv.DictReader((output_dir / "case_summary.csv").open(encoding="utf-8-sig")))
    assert [row["案件ID"] for row in rows] == ["CASE-001"]


def test_scan_options_can_switch_document_density_to_character_unit(tmp_path):
    root = tmp_path / "cases"
    case_001 = root / "CASE-001" / "レビュー結果記録表.xlsx"
    _write_case_workbook(case_001, "CASE-001", "注文登録", "2026-05-01", "2026-05-10")
    output_dir = tmp_path / "out"

    scan_review_root_with_options(
        ReviewScanOptions(
            root=root,
            output_dir=output_dir,
            skip_bazaar=True,
            document_density_unit="characters",
        )
    )

    rows = list(csv.DictReader((output_dir / "phase_metrics.csv").open(encoding="utf-8-sig")))
    external = next(row for row in rows if row["工程"] == "外部仕様書")
    assert external["密度分母"] == "レビュー対象文字数"
    assert float(external["分母値"]) == 5000
    assert float(external["指摘密度"]) == 1 * 1000 / 5000
    assert external["指摘密度単位"] == "件/1000文字"


def test_character_density_unit_reports_missing_characters_instead_of_missing_pages(tmp_path):
    root = tmp_path / "cases"
    case_001 = root / "CASE-001" / "レビュー結果記録表.xlsx"
    _write_case_workbook(case_001, "CASE-001", "注文登録", "2026-05-01", "2026-05-10")

    workbook = Workbook()
    workbook.remove(workbook.active)
    management = workbook.create_sheet("_集計管理")
    for row in [
        ("項目", "値"),
        ("案件ID", "CASE-EMPTY"),
        ("案件名", "分母未入力"),
        ("コード変更ステップ数", 100),
    ]:
        management.append(row)
    sheet = workbook.create_sheet("外部仕様書")
    sheet.append(["指摘項目"])
    sheet.append(HEADERS)
    sheet.append([1, "B", "1章", "指摘", "2026-05-10", "対応済", "", "佐藤", "田中", "不良", "対象", "外部仕様書", "外部仕様書"])
    workbook.save(case_001)
    output_dir = tmp_path / "out"

    scan_review_root_with_options(
        ReviewScanOptions(
            root=root,
            output_dir=output_dir,
            skip_bazaar=True,
            document_density_unit="characters",
        )
    )

    errors = list(csv.DictReader((output_dir / "validation_errors.csv").open(encoding="utf-8-sig")))
    codes = [row["コード"] for row in errors]
    assert "missing_document_characters" in codes
    assert "missing_document_pages" not in codes


def test_gui_settings_round_trip_json(tmp_path):
    settings_path = tmp_path / "settings.json"
    settings = {
        "case_root": "C:/cases",
        "output_dir": "C:/out",
        "start_date": "2026-05-01",
        "end_date": "2026-05-31",
        "skip_bazaar": True,
        "document_density_unit": "characters",
        "included_workbook_paths": ["C:/cases/CASE-001/レビュー結果記録表.xlsx"],
        "excluded_workbook_paths": ["C:/cases/CASE-002/レビュー結果記録表.xlsx"],
        "bazaar_repo_path": "C:/repo",
        "bazaar_from_revision": "100",
        "bazaar_to_revision": "120",
        "before_folder": "C:/before",
        "after_folder": "C:/after",
        "word_document_path": "C:/docs/spec.docx",
        "word_chars_per_page": "1400",
    }

    save_gui_settings(settings_path, settings)

    loaded = load_gui_settings(settings_path)
    assert loaded == settings
