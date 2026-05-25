from __future__ import annotations

import argparse
import csv
from pathlib import Path
import shutil
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))

from bzr_step_count.review_scan import scan_review_root  # noqa: E402


DEFAULT_DEMO_ROOT = ROOT / "samples" / "review_stats_demo"
DEMO_ROOT = DEFAULT_DEMO_ROOT
CASES_ROOT = DEMO_ROOT / "review_cases"
OUTPUT_ROOT = DEMO_ROOT / "aggregate_outputs"
PHASES = ["外部仕様書", "内部仕様書", "コード", "テスト仕様書", "リリース後"]
FINDING_HEADERS = [
    "No",
    "重大度",
    "指摘箇所",
    "指摘内容",
    "対応日時",
    "状況",
    "備考",
    "作業担当者",
    "レビュー担当者",
    "指摘分類",
    "指標対象",
    "検出工程",
    "原因工程",
]


CASES = [
    {
        "case_id": "CASE-001",
        "case_name": "注文登録機能",
        "owner": "品質管理A",
        "reviewer": "レビューリーダー甲",
        "code_steps": 420,
        "pages": {"外部仕様書": 18, "内部仕様書": 12, "テスト仕様書": 9},
        "escaped": 1,
        "findings": {
            "外部仕様書": [
                ["A", "1.2 業務ルール", "取消時の在庫戻し条件が未記載", "不良", "対象", "対応済", "外部仕様書"],
                ["D", "表紙", "版数表記の揺れ", "軽微", "除外", "対応済", "外部仕様書"],
            ],
            "内部仕様書": [
                ["B", "DB更新仕様", "重複登録時のロールバック条件が曖昧", "不良", "対象", "対応済", "内部仕様書"],
                ["C", "画面遷移", "説明文を補足した方がよい", "改善", "対象", "対応保留", "内部仕様書"],
            ],
            "コード": [
                ["B", "src/order_service.py", "数量0の境界値チェックが不足", "不良", "対象", "対応済", "コード"],
                ["D", "src/order_service.py", "コメント内の誤字", "軽微", "除外", "対応済", "コード"],
            ],
            "テスト仕様書": [
                ["C", "異常系ケース", "取消済み注文の再取消ケースが不足", "不良", "対象", "対応済", "テスト仕様書"],
            ],
            "リリース後": [
                ["B", "結合試験", "取消後の通知抑止条件が漏れていた", "不良", "対象", "対応済", "外部仕様書", "後工程"],
            ],
        },
    },
    {
        "case_id": "CASE-002",
        "case_name": "在庫引当改善",
        "owner": "品質管理B",
        "reviewer": "レビューリーダー乙",
        "code_steps": 760,
        "pages": {"外部仕様書": 11, "内部仕様書": 15, "テスト仕様書": 14},
        "escaped": 0,
        "findings": {
            "外部仕様書": [
                ["B", "2.1 引当優先順位", "同一優先度時の並び順が未定義", "不良", "対象", "対応済", "外部仕様書"],
            ],
            "内部仕様書": [
                ["A", "排他制御", "同時更新時の再試行条件が不足", "不良", "対象", "対応済", "内部仕様書"],
                ["C", "ログ仕様", "運用調査用ログ項目の追加提案", "改善", "対象", "対応済", "内部仕様書"],
            ],
            "コード": [
                ["B", "src/stock_allocator.cpp", "ロック解放漏れの可能性", "不良", "対象", "対応済", "コード"],
                ["C", "src/stock_allocator.cpp", "インデントずれ", "軽微", "除外", "対応済", "コード"],
                ["C", "src/stock_allocator.cpp", "戻り値名が仕様と不一致", "不良", "対象", "対応保留", "コード"],
            ],
            "テスト仕様書": [
                ["B", "負荷試験", "同時引当の競合ケースが不足", "不良", "対象", "対応済", "テスト仕様書"],
                ["D", "表記", "表の罫線崩れ", "軽微", "除外", "対応不要", "テスト仕様書"],
            ],
            "リリース後": [],
        },
    },
    {
        "case_id": "CASE-003",
        "case_name": "帳票CSV出力",
        "owner": "品質管理C",
        "reviewer": "レビューリーダー甲",
        "code_steps": 280,
        "pages": {"外部仕様書": 8, "内部仕様書": 9, "テスト仕様書": 7},
        "escaped": 2,
        "findings": {
            "外部仕様書": [
                ["C", "出力項目", "NULL時の空欄/0出力ルールが曖昧", "不良", "対象", "対応済", "外部仕様書"],
                ["D", "本文", "句読点の揺れ", "軽微", "除外", "対応済", "外部仕様書"],
            ],
            "内部仕様書": [
                ["B", "文字コード", "Shift_JIS変換失敗時の扱いが未記載", "不良", "対象", "対応済", "内部仕様書"],
            ],
            "コード": [
                ["A", "src/csv_exporter.cs", "改行を含む項目のクォート処理が不足", "不良", "対象", "対応済", "コード"],
                ["C", "src/csv_exporter.cs", "例外メッセージが利用者向けでない", "改善", "対象", "対応済", "コード"],
            ],
            "テスト仕様書": [
                ["B", "境界値", "最大桁数の出力確認が不足", "不良", "対象", "対応済", "テスト仕様書"],
                ["C", "観点", "空ファイル出力時の期待値を追記", "不良", "対象", "対応済", "テスト仕様書"],
            ],
            "リリース後": [
                ["A", "リリース後障害", "改行を含むCSVが取込先で分割された", "不良", "対象", "対応済", "コード", "リリース後"],
                ["B", "後工程試験", "文字コード変換エラーの再実行手順が不足", "不良", "対象", "対応済", "内部仕様書", "後工程"],
            ],
        },
    },
]


def main() -> int:
    global DEMO_ROOT, CASES_ROOT, OUTPUT_ROOT
    args = _parse_args()
    DEMO_ROOT = Path(args.demo_root).resolve() if args.demo_root else DEFAULT_DEMO_ROOT
    CASES_ROOT = DEMO_ROOT / "review_cases"
    OUTPUT_ROOT = DEMO_ROOT / "aggregate_outputs"

    _reset_generated_tree()
    CASES_ROOT.mkdir(parents=True, exist_ok=True)
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)

    for case in CASES:
        case_dir = CASES_ROOT / case["case_id"]
        case_dir.mkdir()
        _write_case_workbook(case_dir / "レビュー結果記録表.xlsx", case)

    scan_review_root(CASES_ROOT, OUTPUT_ROOT, skip_bazaar=True)
    _write_aggregate_workbook(OUTPUT_ROOT / "集計結果サマリー.xlsx")
    _write_demo_readme()
    print(f"Generated sample input workbooks under: {CASES_ROOT}")
    print(f"Generated aggregate outputs under: {OUTPUT_ROOT}")
    return 0


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--demo-root", help="Output folder for generated sample inputs and reports")
    return parser.parse_args()


def _reset_generated_tree() -> None:
    for path in [CASES_ROOT, OUTPUT_ROOT]:
        if not path.exists():
            continue
        for child in path.iterdir():
            if child.name.startswith("~$"):
                continue
            if child.is_dir():
                shutil.rmtree(child)
            else:
                child.unlink()


def _write_case_workbook(path: Path, case: dict) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_management_sheet(workbook, case)
    for phase in PHASES:
        _write_phase_sheet(workbook, phase, case["findings"].get(phase, []))
    workbook.save(path)


def _write_management_sheet(workbook: Workbook, case: dict) -> None:
    sheet = workbook.create_sheet("_集計管理")
    sheet.append(["項目", "値"])
    rows = [
        ("案件ID", case["case_id"]),
        ("案件名", case["case_name"]),
        ("Bazaarリポジトリパス", ""),
        ("コードfromリビジョン", "100"),
        ("コードtoリビジョン", "120"),
        ("コード変更ステップ数", case["code_steps"]),
        ("外部仕様書ページ数", case["pages"]["外部仕様書"]),
        ("内部仕様書ページ数", case["pages"]["内部仕様書"]),
        ("テスト仕様書ページ数", case["pages"]["テスト仕様書"]),
        (
            "流出不良件数",
            '=COUNTIFS(\'リリース後\'!J:J,"不良",\'リリース後\'!K:K,"対象")'
            '+COUNTIFS(\'リリース後\'!J:J,"不良",\'リリース後\'!K:K,"")',
        ),
        ("Redmine issue", ""),
        ("Redmine URL", ""),
        ("担当者", case["owner"]),
        ("レビュー担当者", case["reviewer"]),
        ("レビュー開始日", "2026-05-01"),
        ("レビュー終了日", "2026-05-20"),
    ]
    for row in rows:
        sheet.append(row)
    _style_header(sheet[1])
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 36


def _write_phase_sheet(workbook: Workbook, phase: str, findings: list[list[str]]) -> None:
    sheet = workbook.create_sheet(phase)
    sheet.append(["レビュー者観点チェックリスト"])
    sheet.append(["No", "内容", "状況", "確認日時"])
    sheet.append([1, f"{phase}の整合性を確認", "確認済", "2026-05-12"])
    sheet.append([2, f"{phase}の変更範囲を確認", "確認済", "2026-05-12"])
    sheet.append([])
    sheet.append(["共通チェックリスト"])
    sheet.append(["No", "内容", "作業者確認日時", "レビュー者確認日時", "状況"])
    sheet.append([1, "成果物の版数と更新日が一致している", "2026-05-10", "2026-05-12", "確認済"])
    sheet.append([2, "レビュー指摘の対応方針が記録されている", "2026-05-10", "2026-05-12", "確認済"])
    sheet.append([])
    sheet.append(["指摘項目"])
    sheet.append(FINDING_HEADERS)
    _style_header(sheet[sheet.max_row])
    for index, finding in enumerate(findings, start=1):
        severity, location, description, classification, target, status, origin = finding[:7]
        detection = finding[7] if len(finding) > 7 else phase
        sheet.append(
            [
                index,
                severity,
                location,
                description,
                "2026-05-15" if status != "対応保留" else "",
                status,
                "",
                "品質管理A" if phase in {"外部仕様書", "内部仕様書"} else "品質管理B",
                "レビューリーダー甲" if phase in {"外部仕様書", "テスト仕様書"} else "レビューリーダー乙",
                classification,
                target,
                detection,
                origin,
            ]
        )
    _add_validations(sheet)
    widths = [6, 8, 22, 44, 14, 12, 18, 14, 18, 12, 10, 14, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width


def _style_header(row) -> None:
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")


def _add_validations(sheet) -> None:
    header_row = 12
    ranges = {
        10: '"不良,改善,軽微,質問,対象外"',
        11: '"対象,除外"',
        12: '"外部仕様書,内部仕様書,コード,テスト仕様書,リリース後,後工程"',
        13: '"外部仕様書,内部仕様書,コード,テスト仕様書,不明"',
    }
    for column, formula in ranges.items():
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        sheet.add_data_validation(validation)
        letter = sheet.cell(row=1, column=column).column_letter
        validation.add(f"{letter}{header_row + 1}:{letter}100")


def _write_demo_readme() -> None:
    text = """# レビュー統計サンプル

このフォルダは `scripts/generate_review_stats_samples.py` で生成したデモデータです。

## 入力例

- `review_cases/CASE-001/レビュー結果記録表.xlsx`: 注文登録機能
- `review_cases/CASE-002/レビュー結果記録表.xlsx`: 在庫引当改善
- `review_cases/CASE-003/レビュー結果記録表.xlsx`: 帳票CSV出力

各Excelには `_集計管理` シートと、外部仕様書・内部仕様書・コード・テスト仕様書・リリース後の5工程シートがあります。
指摘項目には `指摘分類`, `指標対象`, `検出工程`, `原因工程` の入力例を入れています。
`後工程` は既存データ互換の別名として検出工程に残し、集計時は `リリース後` と同じ工程に正規化します。

## 集計成果物

- `aggregate_outputs/case_summary.csv`
- `aggregate_outputs/finding_summary.csv`
- `aggregate_outputs/phase_metrics.csv`
- `aggregate_outputs/phase_summary.csv`
- `aggregate_outputs/owner_summary.csv`
- `aggregate_outputs/reviewer_summary.csv`
- `aggregate_outputs/cross_summary.csv`
- `aggregate_outputs/monthly_report.html`
- `aggregate_outputs/集計結果サマリー.xlsx`
- `aggregate_outputs/review_stats.sqlite`
- `aggregate_outputs/validation_errors.csv`
"""
    (DEMO_ROOT / "README.md").write_text(text, encoding="utf-8")


def _write_aggregate_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, csv_name in [
        ("案件別サマリー", "case_summary.csv"),
        ("工程別指標", "phase_metrics.csv"),
        ("工程横断", "phase_summary.csv"),
        ("作業担当者別", "owner_summary.csv"),
        ("レビュー担当者別", "reviewer_summary.csv"),
        ("横断集計", "cross_summary.csv"),
        ("指摘一覧", "finding_summary.csv"),
        ("品質警告", "validation_errors.csv"),
    ]:
        sheet = workbook.create_sheet(sheet_name)
        csv_path = OUTPUT_ROOT / csv_name
        with csv_path.open(encoding="utf-8-sig", newline="") as handle:
            for row in csv.reader(handle):
                sheet.append(row)
        if sheet.max_row:
            _style_header(sheet[1])
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        for column in range(1, sheet.max_column + 1):
            max_len = 0
            for row in range(1, min(sheet.max_row, 50) + 1):
                value = sheet.cell(row=row, column=column).value
                max_len = max(max_len, len("" if value is None else str(value)))
            sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = min(max(max_len + 2, 10), 42)
    workbook.save(path)


if __name__ == "__main__":
    raise SystemExit(main())
