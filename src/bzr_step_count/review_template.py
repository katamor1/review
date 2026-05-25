from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .review_excel import MANAGEMENT_SHEET, REQUIRED_FINDING_COLUMNS
from .review_models import FINAL_PHASE, PHASE_ORDER, PHASE_SHEET_ALIASES


FINDING_HEADERS = [
    "No",
    "重大度",
    "指摘箇所",
    "指摘内容",
    "対応日時",
    "状況",
    "備考",
    "作業担当者",
    "レビュー担当者",
    "指摘分類",
    "指標対象",
    "検出工程",
    "原因工程",
]


MANAGEMENT_ROWS = [
    ("案件ID", "CASE-001"),
    ("案件名", "機能追加案件名"),
    ("Bazaarリポジトリパス", ""),
    ("コードfromリビジョン", ""),
    ("コードtoリビジョン", ""),
    ("コード変更ステップ数", ""),
    ("外部仕様書ページ数", ""),
    ("内部仕様書ページ数", ""),
    ("テスト仕様書ページ数", ""),
    ("流出不良件数", 0),
    ("Redmine issue", ""),
    ("Redmine URL", ""),
    ("担当者", ""),
    ("レビュー担当者", ""),
    ("レビュー開始日", ""),
    ("レビュー終了日", ""),
]


def upgrade_review_workbook_template(source: str | Path, output: str | Path) -> Path:
    source_path = Path(source)
    output_path = Path(output)
    workbook = load_workbook(source_path)
    _ensure_management_sheet(workbook)
    for phase in PHASE_ORDER:
        sheet = _find_phase_sheet(workbook, phase)
        if sheet is None and phase == FINAL_PHASE:
            sheet = _create_phase_sheet(workbook, phase)
        if sheet is not None:
            _ensure_finding_columns(sheet, phase)
    _ensure_escaped_defects_formula(workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _ensure_management_sheet(workbook: Any) -> None:
    if MANAGEMENT_SHEET in workbook.sheetnames:
        sheet = workbook[MANAGEMENT_SHEET]
    else:
        sheet = workbook.create_sheet(MANAGEMENT_SHEET, 0)

    sheet["A1"] = "項目"
    sheet["B1"] = "値"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")

    existing = {_text(sheet.cell(row=row, column=1).value): row for row in range(2, sheet.max_row + 1)}
    row_index = 2
    for key, default in MANAGEMENT_ROWS:
        row = existing.get(key, row_index)
        sheet.cell(row=row, column=1).value = key
        if sheet.cell(row=row, column=2).value in {None, ""}:
            sheet.cell(row=row, column=2).value = default
        row_index = max(row_index, row + 1)
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 40


def _find_phase_sheet(workbook: Any, phase: str) -> Any | None:
    for sheet_name in PHASE_SHEET_ALIASES[phase]:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    return None


def _create_phase_sheet(workbook: Any, phase: str) -> Any:
    sheet = workbook.create_sheet(phase)
    sheet.append(["レビュー者観点チェックリスト"])
    sheet.append(["No", "内容", "状況", "確認日時"])
    sheet.append([1, f"{phase}で検出した不良・指摘を確認", "", ""])
    sheet.append([2, f"{phase}での対応状況を確認", "", ""])
    sheet.append([])
    sheet.append(["共通チェックリスト"])
    sheet.append(["No", "内容", "作業者確認日時", "レビュー者確認日時", "状況"])
    sheet.append([1, "指摘内容と対応状況が記録されている", "", "", ""])
    sheet.append([2, "指標対象外にする理由が指摘分類で判別できる", "", "", ""])
    sheet.append([])
    sheet.append(["指摘項目"])
    sheet.append(FINDING_HEADERS)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
    widths = [6, 8, 22, 44, 14, 12, 18, 14, 18, 12, 10, 14, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    return sheet


def _ensure_finding_columns(sheet: Any, phase: str) -> None:
    header_row = _find_finding_header(sheet)
    if header_row is None:
        return

    headers = {_text(sheet.cell(row=header_row, column=column).value): column for column in range(1, sheet.max_column + 1)}
    next_column = sheet.max_column + 1
    for header in ["作業担当者", "レビュー担当者", *REQUIRED_FINDING_COLUMNS]:
        if header in headers:
            continue
        sheet.cell(row=header_row, column=next_column).value = header
        sheet.cell(row=header_row, column=next_column).font = Font(bold=True)
        headers[header] = next_column
        next_column += 1

    for row in range(header_row + 1, sheet.max_row + 1):
        if _is_empty_finding_row(sheet, row, headers):
            continue
        if sheet.cell(row=row, column=headers["指摘分類"]).value in {None, ""}:
            sheet.cell(row=row, column=headers["指摘分類"]).value = "不良"
        if sheet.cell(row=row, column=headers["指標対象"]).value in {None, ""}:
            sheet.cell(row=row, column=headers["指標対象"]).value = "対象"
        if sheet.cell(row=row, column=headers["検出工程"]).value in {None, ""}:
            sheet.cell(row=row, column=headers["検出工程"]).value = phase
        if sheet.cell(row=row, column=headers["原因工程"]).value in {None, ""}:
            sheet.cell(row=row, column=headers["原因工程"]).value = "不明"

    start_row = header_row + 1
    _add_validation(sheet, headers["指摘分類"], '"不良,改善,軽微,質問,対象外"', start_row)
    _add_validation(sheet, headers["指標対象"], '"対象,除外"', start_row)
    _add_validation(sheet, headers["検出工程"], '"外部仕様書,内部仕様書,コード,テスト仕様書,リリース後,後工程"', start_row)
    _add_validation(sheet, headers["原因工程"], '"外部仕様書,内部仕様書,コード,テスト仕様書,不明"', start_row)


def _ensure_escaped_defects_formula(workbook: Any) -> None:
    if MANAGEMENT_SHEET not in workbook.sheetnames:
        return
    release_sheet = _find_phase_sheet(workbook, FINAL_PHASE)
    if release_sheet is None:
        return
    header_row = _find_finding_header(release_sheet)
    if header_row is None:
        return
    headers = {
        _text(release_sheet.cell(row=header_row, column=column).value): column
        for column in range(1, release_sheet.max_column + 1)
    }
    classification_column = headers.get("指摘分類")
    target_column = headers.get("指標対象")
    if classification_column is None or target_column is None:
        return

    management_sheet = workbook[MANAGEMENT_SHEET]
    formula_row = None
    for row in range(2, management_sheet.max_row + 1):
        if _text(management_sheet.cell(row=row, column=1).value) == "流出不良件数":
            formula_row = row
            break
    if formula_row is None:
        formula_row = management_sheet.max_row + 1
        management_sheet.cell(row=formula_row, column=1).value = "流出不良件数"

    classification_letter = release_sheet.cell(row=1, column=classification_column).column_letter
    target_letter = release_sheet.cell(row=1, column=target_column).column_letter
    quoted_sheet = release_sheet.title.replace("'", "''")
    management_sheet.cell(row=formula_row, column=2).value = (
        f'=COUNTIFS(\'{quoted_sheet}\'!{classification_letter}:{classification_letter},"不良",'
        f'\'{quoted_sheet}\'!{target_letter}:{target_letter},"対象")'
        f'+COUNTIFS(\'{quoted_sheet}\'!{classification_letter}:{classification_letter},"不良",'
        f'\'{quoted_sheet}\'!{target_letter}:{target_letter},"")'
    )


def _find_finding_header(sheet: Any) -> int | None:
    for row in range(1, sheet.max_row + 1):
        values = {_text(sheet.cell(row=row, column=column).value) for column in range(1, sheet.max_column + 1)}
        if "指摘内容" in values and ("重大度" in values or "指摘箇所" in values):
            return row
    return None


def _is_empty_finding_row(sheet: Any, row: int, headers: dict[str, int]) -> bool:
    for key in ["No", "重大度", "指摘箇所", "指摘内容", "状況", "備考"]:
        column = headers.get(key)
        if column and _text(sheet.cell(row=row, column=column).value):
            return False
    return True


def _add_validation(sheet: Any, column: int, formula: str, start_row: int) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    sheet.add_data_validation(validation)
    column_letter = sheet.cell(row=1, column=column).column_letter
    validation.add(f"{column_letter}{start_row}:{column_letter}500")


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()
