import csv
import json
import sqlite3
import subprocess
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bzr_step_count.review_excel import read_review_workbook
from bzr_step_count.review_metrics import build_review_dataset
from bzr_step_count.review_models import CaseMetadata, FindingRecord, ReviewCase
from bzr_step_count.review_scan import scan_review_root
from bzr_step_count.review_template import upgrade_review_workbook_template


PHASES = ["外部仕様書", "内部仕様書", "コード", "テスト仕様書"]


def _create_review_workbook(path: Path, *, missing_classification: bool = False) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    management = workbook.create_sheet("_集計管理")
    management.append(["項目", "値"])
    rows = [
        ("案件ID", "CASE-001"),
        ("案件名", "注文登録機能"),
        ("Bazaarリポジトリパス", ""),
        ("コードfromリビジョン", "100"),
        ("コードtoリビジョン", "120"),
        ("コード変更ステップ数", 250),
        ("外部仕様書ページ数", 20),
        ("内部仕様書ページ数", 12),
        ("テスト仕様書ページ数", 8),
        ("流出不良件数", 1),
    ]
    for row in rows:
        management.append(row)

    for phase in PHASES:
        sheet = workbook.create_sheet(phase)
        sheet.append(["レビュー者観点チェックリスト"])
        sheet.append(["No", "内容", "状況", "確認日時"])
        sheet.append([1, "観点1", "確認済", "2026-05-22"])
        sheet.append([])
        sheet.append(["指摘項目"])
        headers = ["No", "重大度", "指摘箇所", "指摘内容", "対応日時", "状況", "備考"]
        if not missing_classification:
            headers.extend(["作業担当者", "レビュー担当者", "指摘分類", "指標対象", "検出工程", "原因工程"])
        sheet.append(headers)
        if phase == "外部仕様書":
            sheet.append([1, "A", "1章", "要件漏れ", "2026-05-22", "対応済", "", "佐藤", "田中", "不良", "対象", phase, "外部仕様書"])
            sheet.append([2, "D", "表紙", "誤字", "2026-05-22", "対応済", "", "佐藤", "田中", "軽微", "除外", phase, "外部仕様書"])
        elif phase == "コード":
            sheet.append([1, "B", "src/app.py", "境界値漏れ", "2026-05-22", "対応済", "", "鈴木", "山田", "不良", "対象", phase, "コード"])
            sheet.append([2, "C", "src/app.py", "コメント誤字", "2026-05-22", "対応済", "", "鈴木", "山田", "軽微", "除外", phase, "コード"])
        else:
            sheet.append([1, "C", "本文", "改善提案", "", "対応保留", "", "佐藤", "田中", "改善", "対象", phase, phase])

    workbook.save(path)


def test_read_review_workbook_normalizes_management_and_findings(tmp_path):
    workbook_path = tmp_path / "レビュー結果記録表.xlsx"
    _create_review_workbook(workbook_path)

    case = read_review_workbook(workbook_path)

    assert case.metadata.case_id == "CASE-001"
    assert case.metadata.case_name == "注文登録機能"
    assert case.metadata.phase_pages["外部仕様書"] == 20
    assert case.metadata.code_changed_lines == 250
    assert case.metadata.escaped_defects == 0
    assert {finding.phase for finding in case.findings} == set(PHASES)
    assert any(f.classification == "軽微" and not f.metric_target for f in case.findings)
    assert any(f.work_owner == "佐藤" and f.reviewer == "田中" for f in case.findings)


def test_four_phase_workbook_without_release_after_sheet_is_compatible(tmp_path):
    workbook_path = tmp_path / "レビュー結果記録表.xlsx"
    _create_review_workbook(workbook_path)

    case = read_review_workbook(workbook_path)

    assert "リリース後" not in {finding.phase for finding in case.findings}
    assert case.metadata.escaped_defects == 0
    assert not any(
        error.severity == "error" and error.code == "missing_phase_sheet" and "リリース後" in error.message
        for error in case.validation_errors
    )
    assert any(error.severity == "warning" and error.code == "missing_optional_phase_sheet" for error in case.validation_errors)


def test_release_after_sheet_recomputes_escaped_defects_from_metric_targets(tmp_path):
    workbook_path = tmp_path / "レビュー結果記録表.xlsx"
    _create_review_workbook(workbook_path)

    workbook = load_workbook(workbook_path)
    management = workbook["_集計管理"]
    for row in range(1, management.max_row + 1):
        if management.cell(row=row, column=1).value == "流出不良件数":
            management.cell(row=row, column=2).value = 99
            break
    release_sheet = workbook.create_sheet("リリース後")
    release_sheet.append(["指摘項目"])
    release_sheet.append([
        "No",
        "重大度",
        "指摘箇所",
        "指摘内容",
        "対応日時",
        "状況",
        "備考",
        "作業担当者",
        "レビュー担当者",
        "指摘分類",
        "指標対象",
        "検出工程",
        "原因工程",
    ])
    release_sheet.append([1, "A", "本番障害", "条件Aでエラー", "2026-05-22", "対応済", "", "佐藤", "田中", "不良", "対象", "リリース後", "コード"])
    release_sheet.append([2, "B", "本番障害", "条件Bでエラー", "2026-05-22", "対応済", "", "佐藤", "田中", "不良", "対象", "リリース後", "内部仕様書"])
    release_sheet.append([3, "D", "表示", "文言ゆれ", "2026-05-22", "対応済", "", "佐藤", "田中", "軽微", "対象", "リリース後", "外部仕様書"])
    workbook.save(workbook_path)

    case = read_review_workbook(workbook_path)
    dataset = build_review_dataset([case], validation_errors=case.validation_errors)

    assert case.metadata.escaped_defects == 2
    assert dataset.case_summaries[0].escaped_defects == 2
    assert any(finding.phase == "リリース後" and finding.description == "文言ゆれ" for finding in case.findings)


def test_upgrade_template_adds_release_after_sheet_with_finding_table(tmp_path):
    source_path = tmp_path / "レビュー結果記録表.xlsx"
    output_path = tmp_path / "アップグレード済み.xlsx"
    _create_review_workbook(source_path)

    upgrade_review_workbook_template(source_path, output_path)

    workbook = load_workbook(output_path)
    assert "リリース後" in workbook.sheetnames
    release_sheet = workbook["リリース後"]
    header_values = [cell.value for cell in release_sheet[12]]
    for header in ["指摘内容", "作業担当者", "レビュー担当者", "指摘分類", "指標対象", "検出工程", "原因工程"]:
        assert header in header_values


def test_minor_findings_are_never_metric_targets_and_blank_target_has_no_warning(tmp_path):
    workbook_path = tmp_path / "レビュー結果記録表.xlsx"
    _create_review_workbook(workbook_path)
    workbook = Workbook()

    # Use a focused workbook so the behavior under test is easy to inspect.
    workbook.remove(workbook.active)
    management = workbook.create_sheet("_集計管理")
    for row in [
        ("項目", "値"),
        ("案件ID", "CASE-MINOR"),
        ("案件名", "軽微扱い確認"),
        ("コード変更ステップ数", 100),
        ("外部仕様書ページ数", 10),
        ("内部仕様書ページ数", 10),
        ("テスト仕様書ページ数", 10),
        ("流出不良件数", 0),
    ]:
        management.append(row)

    for phase in PHASES:
        sheet = workbook.create_sheet(phase)
        sheet.append(["指摘項目"])
        sheet.append(["No", "重大度", "指摘箇所", "指摘内容", "対応日時", "状況", "備考", "作業担当者", "レビュー担当者", "指摘分類", "指標対象", "検出工程", "原因工程"])
        if phase == "外部仕様書":
            sheet.append([1, "D", "表紙", "誤字", "2026-05-22", "対応済", "", "佐藤", "田中", "軽微", "対象", phase, phase])
            sheet.append([2, "D", "表紙", "罫線ずれ", "2026-05-22", "対応済", "", "佐藤", "田中", "軽微", "", phase, phase])
            sheet.append([3, "B", "1章", "業務ルール漏れ", "2026-05-22", "対応済", "", "佐藤", "田中", "不良", "対象", phase, phase])
    workbook.save(workbook_path)

    case = read_review_workbook(workbook_path)
    dataset = build_review_dataset([case], validation_errors=case.validation_errors)

    minor_findings = [finding for finding in case.findings if finding.classification == "軽微"]
    assert len(minor_findings) == 2
    assert all(not finding.metric_target for finding in minor_findings)
    assert not any(error.code == "missing_metric_target" for error in case.validation_errors)
    assert dataset.case_summaries[0].display_findings == 1
    assert dataset.case_summaries[0].total_findings == 3


def test_build_review_dataset_calculates_density_removal_and_escape_rates(tmp_path):
    workbook_path = tmp_path / "レビュー結果記録表.xlsx"
    _create_review_workbook(workbook_path)
    case = read_review_workbook(workbook_path)

    dataset = build_review_dataset([case], validation_errors=[])

    summary = dataset.case_summaries[0]
    assert summary.display_findings == 4
    assert summary.metric_findings == 4
    assert summary.minor_findings == 2
    assert summary.defect_findings == 2
    assert summary.escaped_defects == 0
    assert summary.escape_rate == 0

    by_phase = {(row.case_id, row.phase): row for row in dataset.phase_metrics}
    assert by_phase[("CASE-001", "外部仕様書")].finding_density == 1 / 20
    assert by_phase[("CASE-001", "外部仕様書")].finding_density_unit == "件/ページ"
    assert by_phase[("CASE-001", "コード")].finding_density == 1 / 250 * 1000
    assert by_phase[("CASE-001", "コード")].finding_density_unit == "件/KLOC"
    assert by_phase[("CASE-001", "コード")].defect_removal_rate == 1


def test_build_review_dataset_can_use_character_density_for_document_phases(tmp_path):
    workbook_path = tmp_path / "レビュー結果記録表.xlsx"
    _create_review_workbook(workbook_path)
    case = read_review_workbook(workbook_path)
    case.metadata.phase_characters["外部仕様書"] = 5000

    dataset = build_review_dataset([case], validation_errors=[], document_density_unit="characters")

    by_phase = {(row.case_id, row.phase): row for row in dataset.phase_metrics}
    external = by_phase[("CASE-001", "外部仕様書")]
    assert external.denominator_name == "レビュー対象文字数"
    assert external.denominator_value == 5000
    assert external.finding_density == 1 * 1000 / 5000
    assert external.finding_density_unit == "件/1000文字"
    assert external.character_density_per_1000 == external.finding_density
    assert by_phase[("CASE-001", "コード")].finding_density_unit == "件/KLOC"


def test_phase_metrics_use_origin_phase_for_eligible_escape_responsibility():
    case = ReviewCase(
        metadata=CaseMetadata(case_id="CASE-X", case_name="工程責任確認"),
        findings=[
            _finding("外部仕様書", "外部仕様書", "佐藤", "田中"),
            _finding("コード", "外部仕様書", "佐藤", "山田"),
            _finding("テスト仕様書", "コード", "鈴木", "山田"),
            _finding("後工程", "テスト仕様書", "鈴木", "田中"),
        ],
    )

    dataset = build_review_dataset([case], validation_errors=[])

    by_phase = {row.phase: row for row in dataset.phase_metrics}
    assert by_phase["外部仕様書"].eligible_defects == 2
    assert by_phase["外部仕様書"].escaped_from_phase_defects == 1
    assert by_phase["外部仕様書"].escape_rate == 1 / 2
    assert by_phase["コード"].eligible_defects == 3
    assert by_phase["コード"].escaped_from_phase_defects == 1
    assert by_phase["コード"].escape_rate == 1 / 3

    by_axis = {(row.axis, row.key): row for row in dataset.cross_summaries}
    assert by_axis[("工程", "コード")].defect_findings == 1
    assert by_axis[("作業担当者", "佐藤")].defect_findings == 2
    assert by_axis[("レビュー担当者", "山田")].defect_findings == 2
    assert by_axis[("原因工程", "外部仕様書")].defect_findings == 2


def test_scan_review_root_writes_csv_sqlite_html_and_validation_errors(tmp_path):
    case_dir = tmp_path / "CASE-001"
    case_dir.mkdir()
    _create_review_workbook(case_dir / "レビュー結果記録表.xlsx", missing_classification=True)
    output_dir = tmp_path / "out"

    result = scan_review_root(tmp_path, output_dir, skip_bazaar=True)

    assert len(result.cases) == 1
    assert (output_dir / "case_summary.csv").exists()
    assert (output_dir / "finding_summary.csv").exists()
    assert (output_dir / "phase_metrics.csv").exists()
    assert (output_dir / "cross_summary.csv").exists()
    assert (output_dir / "phase_summary.csv").exists()
    assert (output_dir / "owner_summary.csv").exists()
    assert (output_dir / "reviewer_summary.csv").exists()
    assert (output_dir / "monthly_report.html").exists()
    assert (output_dir / "review_stats.sqlite").exists()

    case_summary_header = (output_dir / "case_summary.csv").read_text(encoding="utf-8-sig").splitlines()[0]
    assert "案件ID" in case_summary_header
    assert "表示対象指摘件数" in case_summary_header
    assert "指標対象指摘件数" in case_summary_header
    assert "case_id" not in case_summary_header

    phase_metrics_header = (output_dir / "phase_metrics.csv").read_text(encoding="utf-8-sig").splitlines()[0]
    assert "指摘密度単位" in phase_metrics_header

    html = (output_dir / "monthly_report.html").read_text(encoding="utf-8")
    assert "<th>案件ID</th>" in html
    assert "<th>表示対象指摘件数</th>" in html
    assert "<th>指標対象指摘件数</th>" in html
    assert '<div>軽微指摘</div><div class="value">' not in html
    assert "工程横断サマリー" in html
    assert "作業担当者別サマリー" in html
    assert "レビュー担当者別サマリー" in html
    assert "<th>指摘密度単位</th>" in html
    assert "<th>case_id</th>" not in html

    errors = list(csv.DictReader((output_dir / "validation_errors.csv").open(encoding="utf-8-sig")))
    assert any(row["コード"] == "missing_finding_columns" for row in errors)

    with sqlite3.connect(output_dir / "review_stats.sqlite") as connection:
        count = connection.execute("select count(*) from findings").fetchone()[0]
    assert count == 6

    payload = json.loads(result.to_json())
    assert payload["case_summaries"][0]["case_id"] == "CASE-001"


def test_scan_output_keeps_minor_findings_for_audit_with_metric_target_false(tmp_path):
    case_dir = tmp_path / "CASE-001"
    case_dir.mkdir()
    _create_review_workbook(case_dir / "レビュー結果記録表.xlsx")
    output_dir = tmp_path / "out"

    scan_review_root(tmp_path, output_dir, skip_bazaar=True)

    findings = list(csv.DictReader((output_dir / "finding_summary.csv").open(encoding="utf-8-sig")))
    minor_rows = [row for row in findings if row["指摘分類"] == "軽微"]
    assert minor_rows
    assert all(row["指標対象"] == "false" for row in minor_rows)


def test_scan_without_skip_bazaar_records_detected_changed_lines(monkeypatch, tmp_path):
    case_dir = tmp_path / "CASE-001"
    case_dir.mkdir()
    workbook_path = case_dir / "レビュー結果記録表.xlsx"
    _create_review_workbook(workbook_path)
    repo = tmp_path / "repo"
    repo.mkdir()

    workbook = load_workbook(workbook_path)
    management = workbook["_集計管理"]
    values = {
        "Bazaarリポジトリパス": str(repo),
        "コードfromリビジョン": "r1",
        "コードtoリビジョン": "r2",
        "コード変更ステップ数": "",
    }
    for row in range(1, management.max_row + 1):
        key = management.cell(row=row, column=1).value
        if key in values:
            management.cell(row=row, column=2).value = values[key]
    workbook.save(workbook_path)

    diff = """=== modified file 'src/main.py'
--- src/main.py
+++ src/main.py
@@ -1 +1,2 @@
-old
+new
+added
"""
    monkeypatch.setattr(
        subprocess,
        "run",
        lambda args, **kwargs: subprocess.CompletedProcess(args, 0, stdout=diff, stderr=""),
    )

    result = scan_review_root(tmp_path, tmp_path / "out", skip_bazaar=False, write_outputs=False)

    assert result.cases[0].metadata.code_changed_lines == 3
    assert result.cases[0].metadata.bazaar_detected_changed_lines == 3


def _finding(detection_phase: str, origin_phase: str, owner: str, reviewer: str) -> FindingRecord:
    return FindingRecord(
        case_id="CASE-X",
        case_name="工程責任確認",
        workbook_path="sample.xlsx",
        phase=detection_phase if detection_phase in PHASES else "テスト仕様書",
        sheet="テスト仕様書",
        row=1,
        number="1",
        severity="B",
        location="対象箇所",
        description="指摘内容",
        response_date="2026-05-22",
        status="対応済",
        notes="",
        classification="不良",
        metric_target=True,
        detection_phase=detection_phase,
        origin_phase=origin_phase,
        work_owner=owner,
        reviewer=reviewer,
    )
