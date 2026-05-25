from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .review_models import (
    FINAL_PHASE,
    METRIC_EXCLUDED_CLASSIFICATIONS,
    PHASE_ORDER,
    PHASE_SHEET_ALIASES,
    CaseMetadata,
    FindingRecord,
    ReviewCase,
    ValidationMessage,
)


MANAGEMENT_SHEET = "_集計管理"
REQUIRED_FINDING_COLUMNS = ["指摘分類", "指標対象", "検出工程", "原因工程"]

KEY_ALIASES = {
    "case_id": ["案件ID", "case_id", "Case ID"],
    "case_name": ["案件名", "機能名", "case_name"],
    "bazaar_repo_path": ["Bazaarリポジトリパス", "bazaar_repo_path", "リポジトリパス"],
    "from_revision": ["コードfromリビジョン", "from_revision", "from"],
    "to_revision": ["コードtoリビジョン", "to_revision", "to"],
    "code_changed_lines": ["コード変更ステップ数", "変更ステップ数", "code_changed_lines"],
    "escaped_defects": ["流出不良件数", "後工程不良件数", "escaped_defects"],
    "redmine_issue_id": ["Redmine issue", "Redmine Issue", "redmine_issue_id"],
    "redmine_url": ["Redmine URL", "redmine_url"],
    "owner": ["担当者", "owner"],
    "reviewer": ["レビュー担当者", "レビュア", "reviewer"],
    "review_start": ["レビュー開始日", "review_start"],
    "review_end": ["レビュー終了日", "review_end"],
}

PHASE_PAGE_KEYS = {
    "外部仕様書": ["外部仕様書ページ数", "外部仕様書_ページ数", "external_pages"],
    "内部仕様書": ["内部仕様書ページ数", "内部仕様書_ページ数", "internal_pages"],
    "コード": ["コード対象ページ数", "コードページ数"],
    "テスト仕様書": ["テスト仕様書ページ数", "テスト仕様書_ページ数", "test_pages"],
}

PHASE_CHARACTER_KEYS = {
    "外部仕様書": ["外部仕様書文字数", "external_characters"],
    "内部仕様書": ["内部仕様書文字数", "internal_characters"],
    "テスト仕様書": ["テスト仕様書文字数", "test_characters"],
}


def read_review_workbook(path: str | Path) -> ReviewCase:
    workbook_path = Path(path)
    errors: list[ValidationMessage] = []
    workbook = load_workbook(workbook_path, data_only=True)

    metadata = _read_metadata(workbook, workbook_path, errors)
    findings: list[FindingRecord] = []
    for phase in PHASE_ORDER:
        sheet = _find_phase_sheet(workbook, phase)
        if sheet is None:
            if phase == FINAL_PHASE:
                errors.append(
                    ValidationMessage(
                        "warning",
                        "missing_optional_phase_sheet",
                        f"任意工程シートが見つかりません: {phase}",
                        path=str(workbook_path),
                        case_id=metadata.case_id,
                    )
                )
                continue
            errors.append(
                ValidationMessage(
                    "error",
                    "missing_phase_sheet",
                    f"工程シートが見つかりません: {phase}",
                    path=str(workbook_path),
                    case_id=metadata.case_id,
                )
            )
            continue
        findings.extend(_read_findings(sheet, phase, metadata, errors))

    metadata.escaped_defects = _count_release_after_defects(findings)
    return ReviewCase(metadata=metadata, findings=findings, validation_errors=errors)


def discover_review_workbooks(root: str | Path) -> list[Path]:
    root_path = Path(root)
    if root_path.is_file():
        return [root_path] if _is_review_workbook(root_path) else []
    return sorted(path for path in root_path.rglob("*.xlsx") if _is_review_workbook(path))


def _is_review_workbook(path: Path) -> bool:
    name = path.name
    return not name.startswith("~$") and "レビュー結果記録表" in name


def _read_metadata(workbook: Any, path: Path, errors: list[ValidationMessage]) -> CaseMetadata:
    values: dict[str, Any] = {}
    if MANAGEMENT_SHEET in workbook.sheetnames:
        sheet = workbook[MANAGEMENT_SHEET]
        for row in sheet.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = _normalize_key(row[0])
            if not key:
                continue
            values[key] = row[1] if len(row) > 1 else None
    else:
        errors.append(
            ValidationMessage(
                "error",
                "missing_management_sheet",
                f"{MANAGEMENT_SHEET} シートがありません",
                path=str(path),
            )
        )

    case_id = _as_text(_lookup(values, KEY_ALIASES["case_id"]))
    if not case_id:
        case_id = path.parent.name or path.stem
        errors.append(
            ValidationMessage("error", "missing_case_id", "案件IDが未入力です", path=str(path), case_id=case_id)
        )

    metadata = CaseMetadata(
        case_id=case_id,
        case_name=_as_text(_lookup(values, KEY_ALIASES["case_name"])),
        workbook_path=str(path),
        bazaar_repo_path=_as_text(_lookup(values, KEY_ALIASES["bazaar_repo_path"])),
        from_revision=_as_text(_lookup(values, KEY_ALIASES["from_revision"])),
        to_revision=_as_text(_lookup(values, KEY_ALIASES["to_revision"])),
        code_changed_lines=_as_float(_lookup(values, KEY_ALIASES["code_changed_lines"])),
        escaped_defects=int(_as_float(_lookup(values, KEY_ALIASES["escaped_defects"]))),
        redmine_issue_id=_as_text(_lookup(values, KEY_ALIASES["redmine_issue_id"])),
        redmine_url=_as_text(_lookup(values, KEY_ALIASES["redmine_url"])),
        owner=_as_text(_lookup(values, KEY_ALIASES["owner"])),
        reviewer=_as_text(_lookup(values, KEY_ALIASES["reviewer"])),
        review_start=_as_text(_lookup(values, KEY_ALIASES["review_start"])),
        review_end=_as_text(_lookup(values, KEY_ALIASES["review_end"])),
    )

    metadata.phase_pages = {
        phase: _as_float(_lookup(values, aliases)) for phase, aliases in PHASE_PAGE_KEYS.items()
    }
    metadata.phase_characters = {
        phase: _as_float(_lookup(values, aliases)) for phase, aliases in PHASE_CHARACTER_KEYS.items()
    }

    _validate_denominators(metadata, path, errors)
    return metadata


def _validate_denominators(metadata: CaseMetadata, path: Path, errors: list[ValidationMessage]) -> None:
    for phase in ["外部仕様書", "内部仕様書", "テスト仕様書"]:
        if metadata.phase_pages.get(phase, 0) <= 0:
            errors.append(
                ValidationMessage(
                    "warning",
                    "missing_document_pages",
                    f"{phase}ページ数が未入力です",
                    path=str(path),
                    case_id=metadata.case_id,
                )
            )

    has_bazaar_range = bool(metadata.bazaar_repo_path and metadata.from_revision and metadata.to_revision)
    if metadata.code_changed_lines <= 0 and not has_bazaar_range:
        errors.append(
            ValidationMessage(
                "warning",
                "missing_code_denominator",
                "コード変更ステップ数、またはBazaarリポジトリパス/from/toリビジョンが未入力です",
                path=str(path),
                case_id=metadata.case_id,
            )
        )


def _find_phase_sheet(workbook: Any, phase: str) -> Any | None:
    for name in PHASE_SHEET_ALIASES[phase]:
        if name in workbook.sheetnames:
            return workbook[name]
    return None


def _read_findings(
    sheet: Any,
    phase: str,
    metadata: CaseMetadata,
    errors: list[ValidationMessage],
) -> list[FindingRecord]:
    header_row, header_map = _find_finding_header(sheet)
    if header_row is None:
        errors.append(
            ValidationMessage(
                "error",
                "missing_finding_table",
                "指摘項目の見出し行が見つかりません",
                path=metadata.workbook_path,
                sheet=sheet.title,
                case_id=metadata.case_id,
            )
        )
        return []

    missing = [name for name in REQUIRED_FINDING_COLUMNS if name not in header_map]
    if missing:
        errors.append(
            ValidationMessage(
                "warning",
                "missing_finding_columns",
                "指摘項目に集計用列がありません: " + ", ".join(missing),
                path=metadata.workbook_path,
                sheet=sheet.title,
                row=header_row,
                case_id=metadata.case_id,
            )
        )

    findings: list[FindingRecord] = []
    blank_count = 0
    for row_number in range(header_row + 1, sheet.max_row + 1):
        values = {
            header: sheet.cell(row_number, column=column).value for header, column in header_map.items()
        }
        if _is_blank_finding_row(values):
            blank_count += 1
            if blank_count >= 5:
                break
            continue
        blank_count = 0

        classification = _as_text(values.get("指摘分類")) or "不良"
        target_text = _as_text(values.get("指標対象"))
        metric_target = _parse_metric_target(target_text, classification)

        if "指摘分類" in header_map and not _as_text(values.get("指摘分類")):
            errors.append(
                ValidationMessage(
                    "warning",
                    "missing_classification",
                    "指摘分類が未入力です",
                    path=metadata.workbook_path,
                    sheet=sheet.title,
                    row=row_number,
                    case_id=metadata.case_id,
                )
            )
        if "指標対象" in header_map and not target_text and classification != "軽微":
            errors.append(
                ValidationMessage(
                    "warning",
                    "missing_metric_target",
                    "指標対象が未入力です",
                    path=metadata.workbook_path,
                    sheet=sheet.title,
                    row=row_number,
                    case_id=metadata.case_id,
                )
            )

        findings.append(
            FindingRecord(
                case_id=metadata.case_id,
                case_name=metadata.case_name,
                workbook_path=metadata.workbook_path,
                phase=phase,
                sheet=sheet.title,
                row=row_number,
                number=_as_text(values.get("No")),
                severity=_as_text(values.get("重大度")),
                location=_as_text(values.get("指摘箇所")),
                description=_as_text(values.get("指摘内容")),
                response_date=_as_text(values.get("対応日時")),
                status=_as_text(values.get("状況")),
                notes=_as_text(values.get("備考")),
                classification=classification,
                metric_target=metric_target,
                detection_phase=_as_text(values.get("検出工程")) or phase,
                origin_phase=_as_text(values.get("原因工程")) or "不明",
                work_owner=_first_text(values, ["作業担当者", "対応担当者", "担当者"]) or metadata.owner,
                reviewer=_first_text(values, ["レビュー担当者", "レビュア"]) or metadata.reviewer,
            )
        )
    return findings


def _count_release_after_defects(findings: list[FindingRecord]) -> int:
    return sum(1 for finding in findings if finding.phase == FINAL_PHASE and finding.is_defect)


def _first_text(values: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        text = _as_text(values.get(key))
        if text:
            return text
    return ""


def _find_finding_header(sheet: Any) -> tuple[int | None, dict[str, int]]:
    for row in range(1, sheet.max_row + 1):
        header_map: dict[str, int] = {}
        for column in range(1, sheet.max_column + 1):
            text = _as_text(sheet.cell(row=row, column=column).value)
            if text:
                header_map[text] = column
        if "指摘内容" in header_map and ("重大度" in header_map or "指摘箇所" in header_map):
            return row, header_map
    return None, {}


def _is_blank_finding_row(values: dict[str, Any]) -> bool:
    keys = ["No", "重大度", "指摘箇所", "指摘内容", "状況", "備考"]
    return all(not _as_text(values.get(key)) for key in keys)


def _parse_metric_target(value: str, classification: str) -> bool:
    if classification in METRIC_EXCLUDED_CLASSIFICATIONS:
        return False
    normalized = value.strip().lower()
    if normalized in {"除外", "対象外", "false", "0", "no", "n", "いいえ"}:
        return False
    return True


def _lookup(values: dict[str, Any], aliases: list[str]) -> Any:
    for alias in aliases:
        normalized = _normalize_key(alias)
        if normalized in values:
            return values[normalized]
    return None


def _normalize_key(value: Any) -> str:
    return _as_text(value).replace(" ", "").replace("　", "").lower()


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_float(value: Any) -> float:
    text = _as_text(value).replace(",", "")
    if not text:
        return 0
    try:
        return float(text)
    except ValueError:
        return 0
